from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date
import os
import openpyxl
import models, schemas, database
from database import get_db, engine

# Create tables
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="MSL Engagement Management System")

# CORS configuration from environment variables
cors_origins = os.getenv("CORS_ORIGINS", "*")
if cors_origins == "*":
    allow_origins = ["*"]
else:
    allow_origins = [origin.strip() for origin in cors_origins.split(",")]

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=allow_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== AUTHENTICATION ====================

DEFAULT_PASSWORD = "Pulse@123"

# Predefined users based on user requirements
PREDEFINED_USERS = [
    {"username": "Dr Nikhilesh Andhi", "employee_id": "E10472", "role": "Scientific Officer"},
    {"username": "Dr Chetan Dilip Rao There", "employee_id": "E10656", "role": "Scientific Officer"},
    {"username": "Dr ShivDinesh Dyarapogu", "employee_id": "E10771", "role": "Scientific Officer"},
    {"username": "SUMANGAL GHATAK", "employee_id": "E9250", "role": "Asst General Manager"},
    {"username": "Dr Hasrsh Chaturvedi", "employee_id": "E9999", "role": "Associate Vice President"},
    {"username": "BLuser1", "employee_id": "E1000", "role": "BL"},
    {"username": "BMuser1", "employee_id": "E2000", "role": "BM"},
]

@app.post("/api/seed-users")
def seed_users(db: Session = Depends(get_db)):
    """Seed the users table with predefined users"""
    # Check if users already exist
    existing_count = db.query(models.User).count()
    if existing_count > 0:
        return {"message": f"Users already seeded. Found {existing_count} users."}
    
    # Create users
    for user_data in PREDEFINED_USERS:
        db_user = models.User(
            username=user_data["username"],
            employee_id=user_data["employee_id"],
            role=user_data["role"],
            password=DEFAULT_PASSWORD
        )
        db.add(db_user)
    
    db.commit()
    return {"message": f"Successfully seeded {len(PREDEFINED_USERS)} users"}

@app.post("/api/login", response_model=schemas.LoginResponse)
def login(login_data: schemas.LoginRequest, db: Session = Depends(get_db)):
    """Login using employee_id and password"""
    # Find user by employee_id
    user = db.query(models.User).filter(models.User.employee_id == login_data.employee_id).first()
    
    if not user:
        raise HTTPException(status_code=401, detail="Invalid employee ID or password")
    
    # Simple password check (no hashing as requested)
    if user.password != login_data.password:
        raise HTTPException(status_code=401, detail="Invalid employee ID or password")
    
    return {
        "username": user.username,
        "role": user.role,
        "employee_id": user.employee_id,
        "message": "Login successful"
    }

@app.get("/api/users", response_model=List[schemas.User])
def get_users(db: Session = Depends(get_db)):
    """Get all users"""
    users = db.query(models.User).all()
    return users

# ==================== DOCTORS ====================

@app.get("/api/doctors", response_model=List[schemas.Doctor])
def get_doctors(
    priority_only: bool = Query(False, description="Filter priority doctors only"),
    db: Session = Depends(get_db)
):
    """Get all doctors, optionally filtered by priority"""
    query = db.query(models.Doctor)
    if priority_only:
        query = query.filter(models.Doctor.is_priority_doctor == True)
    doctors = query.order_by(
        models.Doctor.is_priority_doctor.desc(),
        models.Doctor.name
    ).all()
    return doctors

@app.post("/api/doctors", response_model=schemas.Doctor)
def create_doctor(doctor: schemas.DoctorCreate, db: Session = Depends(get_db)):
    """Create a new doctor"""
    db_doctor = models.Doctor(**doctor.dict())
    db.add(db_doctor)
    db.commit()
    db.refresh(db_doctor)
    return db_doctor

@app.get("/api/doctors/{doctor_id}", response_model=schemas.Doctor)
def get_doctor(doctor_id: int, db: Session = Depends(get_db)):
    """Get a specific doctor by ID"""
    doctor = db.query(models.Doctor).filter(models.Doctor.id == doctor_id).first()
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    return doctor

@app.post("/api/doctors/{doctor_id}/duplicate", response_model=schemas.Doctor)
def duplicate_doctor(doctor_id: int, db: Session = Depends(get_db)):
    """Duplicate a doctor"""
    doctor = db.query(models.Doctor).filter(models.Doctor.id == doctor_id).first()
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    # Create a duplicate with "(Copy)" suffix
    new_doctor = models.Doctor(
        name=f"{doctor.name} (Copy)",
        speciality=doctor.speciality,
        therapy_area=doctor.therapy_area,
        is_priority_doctor=doctor.is_priority_doctor,
        division=doctor.division,
        territory=doctor.territory,
        emp_code=doctor.emp_code,
        emp_name=doctor.emp_name,
        region=doctor.region,
        doctor_id_ext=doctor.doctor_id_ext,
        uid_number=doctor.uid_number,
        bm_territory=doctor.bm_territory,
        bl_territory=doctor.bl_territory,
        bh_territory=doctor.bh_territory,
        sbuh_territory=doctor.sbuh_territory,
    )
    db.add(new_doctor)
    db.commit()
    db.refresh(new_doctor)
    return new_doctor

@app.delete("/api/doctors/{doctor_id}")
def delete_doctor(doctor_id: int, db: Session = Depends(get_db)):
    """Delete a doctor"""
    doctor = db.query(models.Doctor).filter(models.Doctor.id == doctor_id).first()
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    db.delete(doctor)
    db.commit()
    return {"message": "Doctor deleted successfully"}

# ==================== REQUESTS ====================

@app.post("/api/requests", response_model=schemas.Request)
def create_request(request: schemas.RequestCreate, db: Session = Depends(get_db)):
    """Create a new MSL engagement request"""
    # Verify doctor exists
    doctor = db.query(models.Doctor).filter(models.Doctor.id == request.doctor_id).first()
    if not doctor:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    db_request = models.Request(**request.dict())
    db.add(db_request)
    db.commit()
    db.refresh(db_request)
    return db_request

@app.get("/api/requests", response_model=List[schemas.RequestSummary])
def get_requests(
    requested_by: Optional[str] = None,
    role: Optional[str] = None,
    user_classification: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get requests with optional filters"""
    query = db.query(
        models.Request,
        models.Doctor.name.label("doctor_name")
    ).join(models.Doctor, models.Request.doctor_id == models.Doctor.id)
    
    if requested_by:
        query = query.filter(models.Request.requested_by == requested_by)
    if user_classification:
        query = query.filter(models.Request.user_classification == user_classification)
    
    # Role-based filtering
    if role in ["BL", "BM"]:
        query = query.filter(models.Request.requested_by_role == role)
    
    results = query.order_by(models.Request.created_at.desc()).all()
    
    request_summaries = []
    for result in results:
        request, doctor_name = result
        summary = schemas.RequestSummary(
            id=request.id,
            doctor_id=request.doctor_id,
            requested_by=request.requested_by,
            requested_by_role=request.requested_by_role,
            therapy_area=request.therapy_area,
            objective=request.objective,
            expected_outcome=request.expected_outcome,
            priority=request.priority,
            user_classification=request.user_classification,
            created_at=request.created_at,
            doctor_name=doctor_name
        )
        request_summaries.append(summary)
    
    return request_summaries

@app.get("/api/requests/{request_id}", response_model=schemas.Request)
def get_request(request_id: int, db: Session = Depends(get_db)):
    """Get a specific request with all details"""
    request = db.query(models.Request).filter(models.Request.id == request_id).first()
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    return request

@app.put("/api/requests/{request_id}/user-classification")
def update_request_user_classification(
    request_id: int,
    user_classification: str,
    db: Session = Depends(get_db)
):
    """Update request user classification (potential or non-potential)"""
    valid_classifications = ["potential", "non-potential"]
    if user_classification not in valid_classifications:
        raise HTTPException(status_code=400, detail="Invalid user classification. Must be 'potential' or 'non-potential'")
    
    request = db.query(models.Request).filter(models.Request.id == request_id).first()
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    request.user_classification = user_classification
    db.commit()
    return {"message": "User classification updated successfully"}

# ==================== DOCTOR INTERACTIONS ====================

@app.post("/api/doctor-interactions", response_model=schemas.DoctorInteraction)
def create_doctor_interaction(
    interaction: schemas.DoctorInteractionCreate,
    db: Session = Depends(get_db)
):
    """Log a doctor interaction"""
    # Verify request exists
    request = db.query(models.Request).filter(models.Request.id == interaction.request_id).first()
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    db_interaction = models.DoctorInteraction(**interaction.dict())
    db.add(db_interaction)
    db.commit()
    db.refresh(db_interaction)
    return db_interaction

@app.get("/api/requests/{request_id}/interactions", response_model=List[schemas.DoctorInteraction])
def get_doctor_interactions(request_id: int, db: Session = Depends(get_db)):
    """Get all doctor interactions for a request"""
    interactions = db.query(models.DoctorInteraction).filter(
        models.DoctorInteraction.request_id == request_id
    ).order_by(models.DoctorInteraction.visit_date.desc()).all()
    return interactions

# ==================== OFFICE ACTIVITIES ====================

@app.post("/api/office-activities", response_model=schemas.OfficeActivity)
def create_office_activity(
    activity: schemas.OfficeActivityCreate,
    db: Session = Depends(get_db)
):
    """Log an office activity"""
    db_activity = models.OfficeActivity(**activity.dict())
    db.add(db_activity)
    db.commit()
    db.refresh(db_activity)
    return db_activity

@app.get("/api/office-activities", response_model=List[schemas.OfficeActivity])
def get_office_activities(msl_username: Optional[str] = None, db: Session = Depends(get_db)):
    """Get all office activities, optionally filtered by MSL"""
    query = db.query(models.OfficeActivity)
    if msl_username:
        query = query.filter(models.OfficeActivity.msl_username == msl_username)
    activities = query.order_by(models.OfficeActivity.activity_date.desc()).all()
    return activities

@app.get("/api/office-activities/users", response_model=List[str])
def get_office_activity_users(db: Session = Depends(get_db)):
    """Get unique list of MSL usernames who have logged activities"""
    users = db.query(models.OfficeActivity.msl_username).distinct().all()
    # Filter out None values and return flat list
    return [u[0] for u in users if u[0]]

# ==================== SUMMARY VIEW ====================

@app.get("/api/requests/{request_id}/logs", response_model=List[schemas.ActivityLog])
def get_request_logs(request_id: int, db: Session = Depends(get_db)):
    """Get all activity logs (interactions + office activities) for a request, sorted by date"""
    # Verify request exists
    request = db.query(models.Request).filter(models.Request.id == request_id).first()
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    logs = []
    
    # Get doctor interactions
    interactions = db.query(models.DoctorInteraction).filter(
        models.DoctorInteraction.request_id == request_id
    ).all()
    
    for interaction in interactions:
        logs.append(schemas.ActivityLog(
            id=interaction.id,
            type="doctor_interaction",
            date=interaction.visit_date,
            title=f"Doctor Visit: {interaction.doctor_name}",
            details=interaction.summary,
            created_at=interaction.created_at
        ))
    
    # Sort by date (latest first)
    logs.sort(key=lambda x: x.date, reverse=True)
    
    return logs

# ==================== SEED DATA ====================

@app.post("/api/seed-doctors")
def seed_doctors(db: Session = Depends(get_db)):
    """Seed doctors from Doctors_Format.xlsx (priority_drs_list sheet)."""
    # Locate the Excel file (one directory up from the backend folder)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(base_dir, "..", "Doctors_Format.xlsx")

    if not os.path.exists(excel_path):
        raise HTTPException(
            status_code=500,
            detail=f"Excel file not found at: {excel_path}"
        )

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        sheet_name = "priority_drs_list_01.09.25"
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to open Excel: {str(e)}")

    # Read header row and map column names to 0-based indices
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    added_count = 0
    skipped_count = 0
    seen_uids: set = set()

    # Pre-load existing uid_numbers to avoid duplicate inserts
    existing_uids = {r[0] for r in db.query(models.Doctor.uid_number).all() if r[0]}

    batch = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dr_name = row[col["Dr_Name"]] if "Dr_Name" in col else None
        uid_raw = row[col["uid_number"]] if "uid_number" in col else None
        uid = str(uid_raw).strip() if uid_raw is not None else None

        # Skip empty rows
        if not dr_name:
            continue

        # Deduplicate on uid_number across this batch and existing DB rows
        if uid and (uid in seen_uids or uid in existing_uids):
            skipped_count += 1
            continue
        if uid:
            seen_uids.add(uid)

        def _get(key, default=None, r=row):
            idx = col.get(key)
            if idx is None:
                return default
            val = r[idx]
            return str(val).strip() if val is not None else default

        doctor = models.Doctor(
            name=str(dr_name).strip(),
            speciality=_get("Speciality"),
            therapy_area=_get("Speciality"),   # mirror for backward compatibility
            is_priority_doctor=True,            # all rows in this sheet are priority doctors
            division=_get("Division"),
            territory=_get("Territory"),
            emp_code=_get("Emp_Code"),
            emp_name=_get("Emp_Name"),
            region=_get("Region"),
            doctor_id_ext=_get("Doctor_ID"),
            uid_number=uid,
            bm_territory=_get("BM_Territory"),
            bl_territory=_get("BL_Territory"),
            bh_territory=_get("BH_Territory"),
            sbuh_territory=_get("SBUH_Territory"),
        )
        batch.append(doctor)

        # Bulk-save in batches of 500 to control memory
        if len(batch) >= 500:
            db.bulk_save_objects(batch)
            db.commit()
            added_count += len(batch)
            batch = []

    # Commit any remaining rows
    if batch:
        db.bulk_save_objects(batch)
        db.commit()
        added_count += len(batch)

    wb.close()
    return {
        "message": f"Seeding complete from sheet '{sheet_name}'",
        "added": added_count,
        "skipped_duplicates": skipped_count,
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)